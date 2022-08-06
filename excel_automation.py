import os
from datetime import datetime
import openpyxl
from openpyxl import Workbook, load_workbook
from string import ascii_uppercase


class ExcelAutomation:

    def __init__(self, excel_file: str):
        self.excel_name = excel_file
        self.wb = load_workbook(excel_file)
        self.ws = self.wb.active

    def update_cell(self, cell: str, data) -> None:
        self.ws[cell].value = data
        self.wb.save(self.excel_name)

    #Updates multiple on same row
    def update_multiple_cells(self, new_data: list):
        new_row = self.ws.max_row+1
        counter = 0
        for char in ascii_uppercase:

            if counter == len(new_data):
                break
            else:
                self.ws[char + str(new_row)].value = new_data[counter]
                print(self.ws[char + str(new_row)].value)
                counter = counter + 1
        self.wb.save(self.excel_name)


"""
wb = load_workbook("excel_blad.xlsx")
ws = wb.active
print(ws["A2"].value)
ws["A2"].value = "Bajs"
wb.save("excel_blad.xlsx")
print(ws["A2"].value)
"""

if __name__ == "__main__":
    #xl = ExcelAutomation("excel_blad.xlsx", 3)
    #xl.update_multiple_cells([1,3,4])

    pass

